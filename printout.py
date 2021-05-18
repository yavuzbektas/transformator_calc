
# Scripted t by Yavuz BEKTAS
# 2020

# ================   SETTINGS     ===================================
import os,datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,Font

import myConfig
config_file = myConfig.read_config()

REPORT_DIR = config_file.get("PATHS", "REPORT_DIR")
IMAGE_DIR = config_file.get("PATHS", "IMAGE_DIR")

global guc, frekans, gauss, sac, p_tel_cu_ag, s_tel_cu_ag, p_tel_al_ag, s_tel_al_ag, karkas,trafo_olcu, sac_agirlik,klemens,ayak,sac_tipi
global a_deg,b_deg,c_deg,d_deg,e_deg,f_deg,trafo_a_deg,trafo_b_deg,trafo_c_deg,trafo_d_deg,trafo_e_deg,trafo_f_deg
global primer_izo_deg,primer_izo_tur,sekonder_izo_deg,sekonder_izo_tur,pri_sek_izo_deg,pri_sek_izo_tur,ekran_izo_deg,ekstra_izo_deg,ekran_sec,ekstra

def wb_save(DIR,name,wb):
    wb.save(DIR+name)
# ========================================================================
trafo_tipi = " Izolasyon Trafosu Mono Faz Hesap Ozeti"
tarih = datetime.datetime.now()
img=[]
primer_kademe=2
sekonder_kademe=5
va_kademe=1
va_altkademe=[2,1,4,6,2,1,2,2,10,1]
primer_gp = {}
primer_group_list=[]
sekonder_group_list=[]
va_group_list_1=[]
va_group_list_2=[]
va_group_list_3=[]
va_group_list_4=[]
va_group_list_5=[]
va_group_list_6=[]
va_group_list_7=[]
va_group_list_8=[]
va_group_list_9=[]
va_group_list_10=[]
va_group_list=[]
va_enabled=False
guc=""
gauss=""
sac= ""
c_sac= ""
don_oran=0
frekans=""
p_tel_al_ag = ""
s_tel_al_ag= ""
p_tel_cu_ag= ""
s_tel_cu_ag= ""
toplam_ag = ""
toplam_cu = ""
karkas = ""
karkas_kod=""
trafo_olcu = ""
sac_agirlik =""
sac_tipi=""
klemens = ""
ayak = ""
a_deg = 0
b_deg = 0
c_deg = 0
d_deg = 0
e_deg = 0
f_deg = 0
primer_izo_deg=0
sekonder_izo_deg = 0
pri_sek_izo_deg = 0
ekran_izo_deg = 0
ekstra_izo_deg = 0
primer_izo_tur = 0
sekonder_izo_tur = 0
pri_sek_izo_tur = 0
ekran_sec = False
ekstra = 0
primer_baglanti=""
sekonder_baglanti=""
izolasyon_karsiligi=0
va_guclist=[]
mlz_listesi=[]
kesit_listesi=[]
apSacaGore=0.0
akim_t=0.0
enduktans=0.0
Sp1=0.0
Lg_oto=0.0
f_oto=0.0
Sp_oto=0.0
Lg_mpl=0.0
f_mpl=0.0
Sp_mpl=0.0
Lg_man=0.0
f_man=0.0
Sp_man=0.0
Ap=0.0
mpl=0.0
karkas_cm_man=0.0
karkas_cm_oto=0.0
bosluk=0.0
karkas_yuk_oto =0.0
# ====================================================
    
def topline(ws,tarih,trafo_tipi=trafo_tipi):
    ws.cell(row=1, column=2, value=trafo_tipi)
    ws.cell(row=1, column=14, value=tarih)
    
    ws.cell(row=3, column=2, value=guc)
    ws.cell(row=5, column=2, value=sac)
    ws.cell(row=4, column=2, value=c_sac)
    ws.cell(row=6, column=2, value=gauss)
    ws.cell(row=7, column=2, value=frekans)

    ws.cell(row=3, column=8, value=p_tel_al_ag)
    ws.cell(row=4, column=8, value=s_tel_al_ag)
    ws.cell(row=5, column=8, value=p_tel_cu_ag)
    ws.cell(row=6, column=8, value=s_tel_cu_ag)
    ws.cell(row=7, column=8, value="Al:"+toplam_ag+" //  Cu:"+toplam_cu +"kg")

    ws.cell(row=3, column=12, value=karkas)
    a=ws.cell(row=4, column=12, value=trafo_olcu)
    ws.cell(row=6, column=12, value=sac_agirlik +"(" +sac_tipi+")")
    ws.cell(row=5, column=12, value=klemens+" - "+ayak)

def bottomline(ws, last_row=20, ekran_sec=ekran_sec, ekstra=ekstra):
    if ekran_izo_deg=="" or ekran_izo_deg==0.0:
        ekran_durum="Yok"

    else:
        ekran_durum = chr(216) + " : "+ str(ekran_izo_deg)
        ws.add_image(Image(IMAGE_DIR +'ekran_symbol.png'), "H" + str(12))
    if ekstra_izo_deg=="" or ekstra_izo_deg==0.0:
        ekstra_durum="Yok"
    else:
        ekstra_durum = "Deger: "+ str(ekstra_izo_deg)
    if primer_izo_tur !=0:
        a1=ws.cell(row=last_row + 2, column=2, value=primer_izo_deg)
        a2=ws.cell(row=last_row + 2, column=6, value=primer_izo_tur)
        a1.alignment = Alignment(horizontal='center')
        a2.alignment = Alignment(horizontal='center')
    else:
        a1=ws.cell(row=last_row + 2, column=2, value="Yok")
        a2=ws.cell(row=last_row + 2, column=6, value="Yok")
        a1.alignment = Alignment(horizontal='center')
        a2.alignment = Alignment(horizontal='center')
    if sekonder_izo_tur != 0:

        a1=ws.cell(row=last_row + 4, column=2, value=sekonder_izo_deg)
        a2=ws.cell(row=last_row + 4, column=6, value=sekonder_izo_tur)
        a1.alignment = Alignment(horizontal='center')
        a2.alignment = Alignment(horizontal='center')
    else:
        a1=ws.cell(row=last_row + 4, column=2, value="Yok")
        a2=ws.cell(row=last_row + 4, column=6, value="Yok")
        a1.alignment = Alignment(horizontal='center')
        a2.alignment = Alignment(horizontal='center')
    if pri_sek_izo_tur != 0:

        a1=ws.cell(row=last_row + 6, column=2, value=pri_sek_izo_deg)
        a2=ws.cell(row=last_row + 6, column=6, value=pri_sek_izo_tur)
        a1.alignment = Alignment(horizontal='center')
        a2.alignment = Alignment(horizontal='center')
    else:
        a1=ws.cell(row=last_row + 6, column=2, value="Yok")
        a2=ws.cell(row=last_row + 6, column=6, value="Yok")
        a1.alignment = Alignment(horizontal='center')
        a2.alignment = Alignment(horizontal='center')
    a1=ws.cell(row=last_row + 2, column=7, value=ekran_durum)
    a2=ws.cell(row=last_row + 4, column=7, value=ekstra_durum)
    a1.alignment = Alignment(horizontal='center')
    a2.alignment = Alignment(horizontal='center')


    ws.cell(row=last_row+9, column=2, value=a_deg)
    ws.cell(row=last_row+9, column=3, value=b_deg)
    ws.cell(row=last_row+9, column=4, value=c_deg)
    ws.cell(row=last_row+9, column=5, value=d_deg)
    ws.cell(row=last_row+9, column=6, value=e_deg)
    ws.cell(row=last_row+9, column=7, value=f_deg)

    ws.cell(row=last_row+12, column=2, value=trafo_a_deg)
    ws.cell(row=last_row+12, column=3, value=trafo_b_deg)
    ws.cell(row=last_row+12, column=4, value=trafo_c_deg)
    ws.cell(row=last_row+12, column=5, value=trafo_d_deg)
    ws.cell(row=last_row+12, column=6, value=trafo_e_deg)
    ws.cell(row=last_row+12, column=7, value=trafo_f_deg)
def merge_cols(ws,start_row=0, start_column=2, end_column=7,last_row=33):
    ws.merge_cells(start_row=last_row+start_row, start_column=start_column, end_row=last_row+start_row, end_column=end_column)
def insert_row(ws, start_row,number):
    for i in range(number):
        
        ws.insert_rows(start_row)
def insert_image(ws,start_row,start_col,kademe,image,img_row_size=6):
    for i in range(kademe-1,-1,-1):
        img.append(Image(image))
        if start_col==3:
            adress="C"+str(start_row+img_row_size*i)
        else:
            adress="H"+str(start_row+img_row_size*i)
        
        ws.add_image(Image(image), adress)
    img.clear()
def find_last_row(ws,primer_kademe,sekonder_kademe,va_kademe,va_altkademe,va_enabled,img_row_size=15):
    sat_number=0
    primer_number=0
    sek_number =0
    va_number=0
    if va_enabled:
        for i in range(0,10):
            if i<=va_kademe-1:
                sat_number =(va_altkademe[i]*img_row_size)+sat_number+4

            else:
                pass

        va_number =sat_number-9
        if primer_kademe*img_row_size<sat_number:
            pass
        else : 
            sat_number=(primer_kademe-1)*img_row_size
    else:
        sat_number=(max(primer_kademe,sekonder_kademe)-1)*img_row_size
    primer_number=(primer_kademe-1)*img_row_size
    sek_number=(sekonder_kademe-1)*img_row_size

    return (sat_number,primer_number,sek_number,va_number)
def insert_kademe_value(ws,start_row,kademe_name,kademe,values,image=(IMAGE_DIR +'b4.png'),alt_kademe=0,img_row_size=15):
    if kademe_name=="sekonder":
        start_col=9
        adress="N"
    elif kademe_name=="primer":
        start_col=1
        adress = "F"
    elif kademe_name=="va":
        start_col = 9
        adress = "N"
    else:
        return False
    for i in range(kademe-1,-1,-1):
        a1 =ws.cell(row=start_row-i*img_row_size, column=start_col+6, value=str(values[i]["voltaj"])+" V" + " ( " + str(round(values[i]["spir2"]*don_oran,4))+"V )" )

        a2 =ws.cell(row=start_row+9-i*img_row_size, column=start_col+6, value=str(values[i]["spir2"])+" sp")
        a3 = ws.cell(row=start_row + 2 - i * img_row_size, column=start_col,value="Tel :")
        a4 = ws.cell(row=start_row + 3 - i * img_row_size, column=start_col,value="Cap :")
        a5 = ws.cell(row=start_row + 4 - i * img_row_size, column=start_col, value="Kesit :")
        a6 = ws.cell(row=start_row + 5 - i * img_row_size, column=start_col, value="Akim :")
        a7 = ws.cell(row=start_row + 6 - i * img_row_size, column=start_col,value="Kat :")
        a8 = ws.cell(row=start_row + 7 - i * img_row_size, column=start_col,value="Uzunluk :")
        a9 = ws.cell(row=start_row + 8 - i * img_row_size, column=start_col+2, value="-- Teller --")
        a10 = ws.cell(row=start_row + 9 - i * img_row_size, column=start_col,value="Yuv. Tel :")
        a11 = ws.cell(row=start_row + 10 - i * img_row_size, column=start_col, value="Ytel Agr :")
        a12 = ws.cell(row=start_row + 11 - i * img_row_size, column=start_col, value="Kare Tel :")
        a13= ws.cell(row=start_row + 12 - i * img_row_size, column=start_col, value="Ktel Agr :")
        a14 = ws.cell(row=start_row + 13 - i * img_row_size, column=start_col, value="Folyo :")
        a15 = ws.cell(row=start_row + 14 - i * img_row_size, column=start_col, value="Folyo Agr :")
        a16 = ws.cell(row=start_row + 15 - i * img_row_size, column=start_col, value="Kapton :")
        a17 = ws.cell(row=start_row+1 - i * img_row_size, column=start_col, value=f"{i+1} NOLU KADEME:")

        a17.alignment = Alignment(horizontal='center')
        a17.font = Font(size=12, bold=True,color="FF0000")
        ws.merge_cells(start_row=start_row +1 - i * img_row_size, start_column=start_col, end_column=4 + start_col,
                       end_row=start_row+1 - i * img_row_size)
        if kademe_name=="va":

            a18 = ws.cell(row=start_row + 8 - i * img_row_size, column=start_col + 6, value=f"Guc Degeri : {va_guclist[alt_kademe]}")

        a20 = ws.cell(row=start_row + 2 - i * img_row_size, column=start_col+1, value=values[i]["teltipi"])
        a21 = ws.cell(row=start_row + 3 - i * img_row_size, column=start_col+1, value=values[i]["cap2"])
        a22 = ws.cell(row=start_row + 4 - i * img_row_size, column=start_col+1, value=values[i]["kesit2"])
        a23 = ws.cell(row=start_row + 5 - i * img_row_size, column=start_col+1, value=values[i]["akim2"])
        a25 = ws.cell(row=start_row + 6 - i * img_row_size, column=start_col+1, value=int(values[i]["kat"]))
        a26 = ws.cell(row=start_row + 7 - i * img_row_size, column=start_col+1, value=values[i]["tel_uzunluk"])
        a27 = ws.cell(row=start_row + 8 - i * img_row_size, column=start_col+1, value="")
        if values[i]["mancap_1"]!=0.0:
            a28 = ws.cell(row=start_row + 9 - i * img_row_size, column=start_col+1, value=values[i]["mancap_1"])
            a29 = ws.cell(row=start_row + 10 - i * img_row_size, column=start_col+1, value=values[i]["agr_tel_1"])
            a28.alignment = Alignment(horizontal='center')
            a29.alignment = Alignment(horizontal='center')
            a28.font = Font(size=9)
            a29.font = Font(size=9)
        if values[i]["karetel11"] != 0.0:
            a30 = ws.cell(row=start_row + 11 - i * img_row_size, column=start_col+1, value=str(values[i]["karetel11"])+"x"+str(values[i]["karetel12"]))
            a31 = ws.cell(row=start_row + 12 - i * img_row_size, column=start_col+1, value=values[i]["agr_karetel_1"])
            a30.alignment = Alignment(horizontal='center')
            a31.alignment = Alignment(horizontal='center')
            a30.font = Font(size=9)
            a31.font = Font(size=9)
        if values[i]["folyotel11"] != 0.0:
            a32 = ws.cell(row=start_row + 13 - i * img_row_size, column=start_col+1, value=str(values[i]["folyotel11"])+"x"+str(values[i]["folyotel12"]))
            a33 = ws.cell(row=start_row + 14 - i * img_row_size, column=start_col+1, value=values[i]["agr_folyo_1"])
            a32.alignment = Alignment(horizontal='center')
            a33.alignment = Alignment(horizontal='center')
            a32.font = Font(size=9)
            a33.font = Font(size=9)
        if values[i]["kapton1"] != 0.0:
            a34 = ws.cell(row=start_row + 15 - i * img_row_size, column=start_col+1, value=values[i]["kapton1"])
            a34.alignment = Alignment(horizontal='center')
        if values[i]["mancap_2"] != 0.0:
            a35 = ws.cell(row=start_row + 9 - i * img_row_size, column=start_col + 2, value=values[i]["mancap_2"])
            a36 = ws.cell(row=start_row + 10 - i * img_row_size, column=start_col + 2, value=values[i]["agr_tel_2"])
            a35.alignment = Alignment(horizontal='center')
            a36.alignment = Alignment(horizontal='center')
            a35.font = Font(size=9)
            a36.font = Font(size=9)
        if values[i]["karetel21"] != 0.0:
            a37 = ws.cell(row=start_row + 11 - i * img_row_size, column=start_col + 2, value=str(values[i]["karetel21"])+"x"+str(values[i]["karetel22"]))
            a38 = ws.cell(row=start_row + 12 - i * img_row_size, column=start_col + 2, value=values[i]["agr_karetel_2"])
            a37.alignment = Alignment(horizontal='center')
            a38.alignment = Alignment(horizontal='center')
            a37.font = Font(size=9)
            a38.font = Font(size=9)
        if values[i]["folyotel21"] != 0.0:
            a39 = ws.cell(row=start_row + 13 - i * img_row_size, column=start_col + 2, value=str(values[i]["folyotel21"])+"x"+str(values[i]["folyotel22"]))
            a40 = ws.cell(row=start_row + 14 - i * img_row_size, column=start_col + 2, value=values[i]["agr_folyo_2"])
            a39.alignment = Alignment(horizontal='center')
            a40.alignment = Alignment(horizontal='center')
            a39.font = Font(size=9)
            a40.font = Font(size=9)
        if values[i]["kapton2"] != 0.0:
            a41 = ws.cell(row=start_row + 15 - i * img_row_size, column=start_col + 2, value=values[i]["kapton2"])
            a41.alignment = Alignment(horizontal='center')
            a41.font = Font(size=9)
        if values[i]["mancap_3"] != 0.0:
            a42 = ws.cell(row=start_row + 9 - i * img_row_size, column=start_col + 3, value=values[i]["mancap_3"])
            a43 = ws.cell(row=start_row + 10 - i * img_row_size, column=start_col + 3, value=values[i]["agr_tel_3"])
            a42.alignment = Alignment(horizontal='center')
            a43.alignment = Alignment(horizontal='center')
            a42.font = Font(size=9)
            a43.font = Font(size=9)
        if values[i]["karetel31"] != 0.0:
            a44 = ws.cell(row=start_row + 11 - i * img_row_size, column=start_col + 3, value=str(values[i]["karetel31"])+"x"+str(values[i]["karetel32"]))
            a45 = ws.cell(row=start_row + 12 - i * img_row_size, column=start_col + 3, value=values[i]["agr_karetel_3"])
            a44.alignment = Alignment(horizontal='center')
            a45.alignment = Alignment(horizontal='center')
            a44.font = Font(size=9)
            a45.font = Font(size=9)
        if values[i]["folyotel31"] != 0.0:
            a46 = ws.cell(row=start_row + 13 - i * img_row_size, column=start_col + 3, value=str(values[i]["folyotel31"])+"x"+str(values[i]["folyotel32"]))
            a47 = ws.cell(row=start_row + 14 - i * img_row_size, column=start_col + 3, value=values[i]["agr_folyo_3"])
            a46.alignment = Alignment(horizontal='center')
            a47.alignment = Alignment(horizontal='center')
            a46.font = Font(size=9)
            a47.font = Font(size=9)
        if values[i]["kapton3"] != 0.0:
            a48 = ws.cell(row=start_row + 15 - i * img_row_size, column=start_col + 3, value=values[i]["kapton3"])
            a48.alignment = Alignment(horizontal='center')
            a48.font = Font(size=9)
        if values[i]["mancap_4"] != 0.0:
            a49 = ws.cell(row=start_row + 9 - i * img_row_size, column=start_col + 4, value=values[i]["mancap_4"])
            a50 = ws.cell(row=start_row + 10 - i * img_row_size, column=start_col + 4, value=values[i]["agr_tel_4"])
            a49.alignment = Alignment(horizontal='center')
            a50.alignment = Alignment(horizontal='center')
            a49.font = Font(size=9)
            a50.font = Font(size=9)
        if values[i]["karetel41"] != 0.0:
            a51 = ws.cell(row=start_row + 11 - i * img_row_size, column=start_col + 4, value=str(values[i]["karetel41"])+"x"+str(values[i]["karetel42"]))
            a52 = ws.cell(row=start_row + 12 - i * img_row_size, column=start_col + 4, value=values[i]["agr_karetel_4"])
            a51.alignment = Alignment(horizontal='center')
            a52.alignment = Alignment(horizontal='center')
            a51.font = Font(size=9)
            a52.font = Font(size=9)
        if values[i]["folyotel41"] != 0.0:
            a53 = ws.cell(row=start_row + 13 - i * img_row_size, column=start_col + 4, value=str(values[i]["folyotel41"])+"x"+str(values[i]["folyotel42"]))
            a54 = ws.cell(row=start_row + 14 - i * img_row_size, column=start_col + 4, value=values[i]["agr_folyo_4"])
            a53.alignment = Alignment(horizontal='center')
            a54.alignment = Alignment(horizontal='center')
            a53.font = Font(size=9)
            a54.font = Font(size=9)
        if values[i]["kapton4"] != 0.0:
            a55 = ws.cell(row=start_row + 15 - i * img_row_size, column=start_col + 4, value=values[i]["kapton4"])
            a55.alignment = Alignment(horizontal='center')
            a55.font = Font(size=9)
        a1.alignment = Alignment(horizontal='left')
        a2.alignment = Alignment(horizontal='left')
        a3.alignment = Alignment(horizontal='right')
        a4.alignment = Alignment(horizontal='right')
        a5.alignment = Alignment(horizontal='right')
        a6.alignment = Alignment(horizontal='right')
        a7.alignment = Alignment(horizontal='right')
        a8.alignment = Alignment(horizontal='right')
        a9.alignment = Alignment(horizontal='right')
        a10.alignment = Alignment(horizontal='right')
        a11.alignment = Alignment(horizontal='right')
        a12.alignment = Alignment(horizontal='right')
        a13.alignment = Alignment(horizontal='right')
        a14.alignment = Alignment(horizontal='right')
        a15.alignment = Alignment(horizontal='right')
        a16.alignment = Alignment(horizontal='right')
        a3.font = Font(size=12, bold=True)
        a4.font = Font(size=12, bold=True)
        a5.font = Font(size=12, bold=True)
        a6.font = Font(size=12, bold=True)
        a7.font = Font(size=12, bold=True)
        a8.font = Font(size=12, bold=True)
        a9.font = Font(size=12, bold=True)
        a10.font = Font(size=12, bold=True)
        a11.font = Font(size=12, bold=True)
        a12.font = Font(size=12, bold=True)
        a13.font = Font(size=12, bold=True)
        a14.font = Font(size=12, bold=True)
        a15.font = Font(size=12, bold=True)
        a16.font = Font(size=12, bold=True)

        a20.alignment = Alignment(horizontal='center')
        a21.alignment = Alignment(horizontal='center')
        a22.alignment = Alignment(horizontal='center')
        a23.alignment = Alignment(horizontal='center')
        a25.alignment = Alignment(horizontal='center')
        a26.alignment = Alignment(horizontal='center')
        a27.alignment = Alignment(horizontal='center')




        ws.add_image(Image(image), adress + str(start_row+1-i*img_row_size))
def insert_image_trafo(ws,last_row,image):
    ws.add_image(image, "B"+str(last_row+2))
def mlz_listesi_olustur(ws2,start_row_list):
    start_row_list=3
    for key,value in mlz_listesi.items():

        a20 = ws2.cell(row=start_row_list, column=2, value=key)
        a21 = ws2.cell(row=start_row_list, column=4, value=value)
        start_row_list+=1
    a20 = ws2.cell(row=start_row_list, column=2, value=karkas_kod)  # karkas kodu
    a21 = ws2.cell(row=start_row_list, column=3, value=karkas) # karkas olculeri
    a22 = ws2.cell(row=start_row_list+1, column=2, value=sac_tipi)  # sac tipi,
    a23 = ws2.cell(row=start_row_list+1, column=3, value=sac)  # sac olcu
    a24 = ws2.cell(row=start_row_list+1, column=4, value=sac_agirlik) # sac agirliği
    start_row_list=3
    for kesit in kesit_listesi:

        a20 = ws2.cell(row=start_row_list, column=3, value=kesit)
        
        start_row_list+=1
def rightSide_values(ws, last_row=9):
    
    a1=ws.cell(row=last_row + 1, column=12, value=bosluk)
    a2=ws.cell(row=last_row + 2, column=12, value=enduktans)
    a3=ws.cell(row=last_row + 3, column=12, value=akim_t)
    a4=ws.cell(row=last_row + 4, column=12, value=apSacaGore)
    a5=ws.cell(row=last_row + 5, column=12, value=Ap)
    a6=ws.cell(row=last_row + 6, column=12, value=mpl)
    a7=ws.cell(row=last_row + 7, column=12, value=karkas_cm_oto)
    a8=ws.cell(row=last_row + 8, column=12, value=karkas_cm_man)
    a9=ws.cell(row=last_row + 9, column=12, value=Sp1)
    a10=ws.cell(row=last_row + 10, column=12, value=Lg_oto)
    a11=ws.cell(row=last_row + 11, column=12, value=f_oto)
    a12=ws.cell(row=last_row + 12, column=12, value=Sp_oto)
    a13=ws.cell(row=last_row + 13, column=12, value=Lg_mpl)
    a14=ws.cell(row=last_row + 14, column=12, value=f_mpl)
    a15=ws.cell(row=last_row + 15, column=12, value=Sp_mpl)
    a16=ws.cell(row=last_row + 16, column=12, value=Lg_man)
    a17=ws.cell(row=last_row + 17, column=12, value=f_man)
    a18=ws.cell(row=last_row + 18, column=12, value=Sp_man)
    a2.alignment = Alignment(horizontal='center')   
def izolasyon_mono_printout():
    wb = load_workbook(REPORT_DIR+ 'demo.xlsx')
    ws=wb.active
    topline(ws=ws,tarih=datetime.datetime.now(),trafo_tipi=trafo_tipi)
    
    bottomline(ws=ws,last_row=30)
    max_number,primer_number,sek_number,va_number=find_last_row(ws=ws,primer_kademe=primer_kademe,sekonder_kademe=sekonder_kademe,va_kademe=va_kademe,va_altkademe=va_altkademe,va_enabled=va_enabled)

    insert_row(ws=ws,start_row=30,number=max_number+4)
    ws.merge_cells(start_row=max_number+4+31, start_column=2, end_column=5, end_row=max_number+4+31)
    ws.merge_cells(start_row=max_number+4+32, start_column=2, end_column=5, end_row=max_number+4+32)
    ws.merge_cells(start_row=max_number+4+33, start_column=2, end_column=5, end_row=max_number+4+33)
    ws.merge_cells(start_row=max_number+4+34, start_column=2, end_column=5, end_row=max_number+4+34)
    ws.merge_cells(start_row=max_number+4+35, start_column=2, end_column=5, end_row=max_number+4+35)
    ws.merge_cells(start_row=max_number+4+36, start_column=2, end_column=5, end_row=max_number+4+36)
    start_row=va_number
    if va_enabled:
        for i in range (0,va_kademe):
            if i<=va_kademe-1:
                insert_kademe_value(ws=ws,start_row=start_row,kademe_name="va",kademe=va_altkademe[i],values=va_group_list[i],alt_kademe=i)
                #insert_image(ws=ws,start_row=10+start_row+i*2,start_col=6,kademe=va_altkademe[i],image=IMAGE_DIR +'b3.png')
                start_row=start_row- va_altkademe[i]*(15)-4
            else:
                pass

    else:
        insert_kademe_value(ws=ws,start_row=10+sek_number,kademe_name="sekonder",kademe=sekonder_kademe,values=sekonder_group_list)
        #insert_image(ws=ws,start_row=11+sek_number,start_col=6,kademe=sekonder_kademe,image=IMAGE_DIR +'b3.png')

    insert_kademe_value(ws=ws,start_row=10+primer_number,kademe_name="primer",kademe=primer_kademe,values=primer_group_list)
    #insert_image(ws=ws,start_row=11,start_col=3,kademe=primer_kademe,image=IMAGE_DIR +'b3.png')
    insert_image_trafo(ws=ws,last_row=max_number+46,image=Image(IMAGE_DIR +"o_m.png"))
    
    ws2=wb.get_sheet_by_name("M_Listesi")
    start_row_list=3
    mlz_listesi_olustur(ws2,start_row_list)
    try:
        wb_save(DIR=REPORT_DIR,name='izolasyon_mono.xlsx',wb=wb)
        file = REPORT_DIR+"izolasyon_mono.xlsx"
        os.startfile(file)
    except Exception as err:
        print(err)
def izolasyon_trifaz_printout():
    wb = load_workbook(REPORT_DIR+ 'demo.xlsx')
    ws=wb.active
    topline(ws=ws,tarih=datetime.datetime.now(),trafo_tipi=trafo_tipi)
    
    bottomline(ws=ws,last_row=30)
    max_number,primer_number,sek_number,va_number=find_last_row(ws=ws,primer_kademe=primer_kademe,sekonder_kademe=sekonder_kademe,va_kademe=va_kademe,va_altkademe=va_altkademe,va_enabled=va_enabled)

    insert_row(ws=ws,start_row=30,number=max_number+4)
    ws.merge_cells(start_row=max_number+4+31, start_column=2, end_column=5, end_row=max_number+4+31)
    ws.merge_cells(start_row=max_number+4+32, start_column=2, end_column=5, end_row=max_number+4+32)
    ws.merge_cells(start_row=max_number+4+33, start_column=2, end_column=5, end_row=max_number+4+33)
    ws.merge_cells(start_row=max_number+4+34, start_column=2, end_column=5, end_row=max_number+4+34)
    ws.merge_cells(start_row=max_number+4+35, start_column=2, end_column=5, end_row=max_number+4+35)
    ws.merge_cells(start_row=max_number+4+36, start_column=2, end_column=5, end_row=max_number+4+36)
    start_row=va_number
    
    insert_kademe_value(ws=ws,start_row=10+sek_number,kademe_name="sekonder",kademe=sekonder_kademe,values=sekonder_group_list)
    
    insert_kademe_value(ws=ws,start_row=10+primer_number,kademe_name="primer",kademe=primer_kademe,values=primer_group_list)
    
    insert_image_trafo(ws=ws,last_row=max_number+46,image=Image(IMAGE_DIR +"o_t.png"))
    
    if primer_baglanti=="Yıldız":
        ws.add_image(Image(IMAGE_DIR + 'star.png'), "G" + str(12))
    elif primer_baglanti=="Üçgen":
        ws.add_image(Image(IMAGE_DIR + 'delta.png'), "G" + str(12))
    else:
        print("Bağlantı seçimi hatalı : ",primer_baglanti)
    if sekonder_baglanti=="Yıldız":

        ws.add_image(Image(IMAGE_DIR + 'star.png'), "O" + str(12))
    elif sekonder_baglanti=="Üçgen":
        ws.add_image(Image(IMAGE_DIR + 'delta.png'), "O" + str(12))
    else:
        pass
    ws2=wb.get_sheet_by_name("M_Listesi")
    start_row_list=3
    mlz_listesi_olustur(ws2,start_row_list)
    try:
        wb_save(DIR=REPORT_DIR,name='izolasyon_trifaz.xlsx',wb=wb)
        file = REPORT_DIR+"izolasyon_trifaz.xlsx"
        os.startfile(file)
    except Exception as err:
        print(err)

        print(err)
def ototrafo_trifaz_printout(izole_gucu):
    wb = load_workbook(REPORT_DIR+ 'demo.xlsx')
    ws=wb.active
    topline(ws=ws,tarih=datetime.datetime.now(),trafo_tipi=trafo_tipi)
    
    bottomline(ws=ws,last_row=30)
    max_number,primer_number,sek_number,va_number=find_last_row(ws=ws,primer_kademe=primer_kademe,sekonder_kademe=0,va_kademe=0,va_altkademe=0,va_enabled=False)

    insert_row(ws=ws,start_row=30,number=max_number+4)
    ws.merge_cells(start_row=max_number+4+31, start_column=2, end_column=5, end_row=max_number+4+31)
    ws.merge_cells(start_row=max_number+4+32, start_column=2, end_column=5, end_row=max_number+4+32)
    ws.merge_cells(start_row=max_number+4+33, start_column=2, end_column=5, end_row=max_number+4+33)
    ws.merge_cells(start_row=max_number+4+34, start_column=2, end_column=5, end_row=max_number+4+34)
    ws.merge_cells(start_row=max_number+4+35, start_column=2, end_column=5, end_row=max_number+4+35)
    ws.merge_cells(start_row=max_number+4+36, start_column=2, end_column=5, end_row=max_number+4+36)
    start_row=0
    
   
    
    insert_kademe_value(ws=ws,start_row=10+primer_number,kademe_name="primer",kademe=primer_kademe,values=primer_group_list)
    if izole_gucu>0:
        ws.cell(row=max_number+31, column=2, value="İzolasyon Trafosu Karşılığı " + str(izole_gucu)+"'VA dir")
        ws.merge_cells(start_row=max_number+31, start_column=2, end_column=12, end_row=max_number+31)
    insert_image_trafo(ws=ws,last_row=max_number+46,image=Image(IMAGE_DIR +"o_t.png"))
    
    
    
    ws2=wb.get_sheet_by_name("M_Listesi")
    start_row_list=3
    mlz_listesi_olustur(ws2,start_row_list)
    
    try:
        wb_save(DIR=REPORT_DIR,name='oto_trifaz.xlsx',wb=wb)
        file = REPORT_DIR+"oto_trifaz.xlsx"
        os.startfile(file)
    except Exception as err:
        print(err)   
def ototrafo_monofaz_printout():
    wb = load_workbook(REPORT_DIR+ 'demo.xlsx')
    ws=wb.active
    topline(ws=ws,tarih=datetime.datetime.now(),trafo_tipi=trafo_tipi)
    
    bottomline(ws=ws,last_row=30)
    max_number,primer_number,sek_number,va_number=find_last_row(ws=ws,primer_kademe=primer_kademe,sekonder_kademe=0,va_kademe=0,va_altkademe=0,va_enabled=False)

    insert_row(ws=ws,start_row=30,number=max_number+4)
    ws.merge_cells(start_row=max_number+4+31, start_column=2, end_column=5, end_row=max_number+4+31)
    ws.merge_cells(start_row=max_number+4+32, start_column=2, end_column=5, end_row=max_number+4+32)
    ws.merge_cells(start_row=max_number+4+33, start_column=2, end_column=5, end_row=max_number+4+33)
    ws.merge_cells(start_row=max_number+4+34, start_column=2, end_column=5, end_row=max_number+4+34)
    ws.merge_cells(start_row=max_number+4+35, start_column=2, end_column=5, end_row=max_number+4+35)
    ws.merge_cells(start_row=max_number+4+36, start_column=2, end_column=5, end_row=max_number+4+36)
    start_row=0
    
   
    
    insert_kademe_value(ws=ws,start_row=10+primer_number,kademe_name="primer",kademe=primer_kademe,values=primer_group_list)
    
    insert_image_trafo(ws=ws,last_row=max_number+46,image=Image(IMAGE_DIR +"o_t.png"))
    
    
    
    ws2=wb.get_sheet_by_name("M_Listesi")
    start_row_list=3
    mlz_listesi_olustur(ws2,start_row_list)
    
    try:
        wb_save(DIR=REPORT_DIR,name='oto_monofaz.xlsx',wb=wb)
        file = REPORT_DIR+"oto_monofaz.xlsx"
        os.startfile(file)
    except Exception as err:
        print(err)
def izolasyon_monofazUI_printout():
    wb = load_workbook(REPORT_DIR+ 'demo.xlsx')
    ws=wb.active
    topline(ws=ws,tarih=datetime.datetime.now(),trafo_tipi=trafo_tipi)
    
    bottomline(ws=ws,last_row=30)
    max_number,primer_number,sek_number,va_number=find_last_row(ws=ws,primer_kademe=primer_kademe,sekonder_kademe=sekonder_kademe,va_kademe=va_kademe,va_altkademe=va_altkademe,va_enabled=va_enabled)

    insert_row(ws=ws,start_row=30,number=max_number+4)
    ws.merge_cells(start_row=max_number+4+31, start_column=2, end_column=5, end_row=max_number+4+31)
    ws.merge_cells(start_row=max_number+4+32, start_column=2, end_column=5, end_row=max_number+4+32)
    ws.merge_cells(start_row=max_number+4+33, start_column=2, end_column=5, end_row=max_number+4+33)
    ws.merge_cells(start_row=max_number+4+34, start_column=2, end_column=5, end_row=max_number+4+34)
    ws.merge_cells(start_row=max_number+4+35, start_column=2, end_column=5, end_row=max_number+4+35)
    ws.merge_cells(start_row=max_number+4+36, start_column=2, end_column=5, end_row=max_number+4+36)
    start_row=va_number
    
    insert_kademe_value(ws=ws,start_row=10+sek_number,kademe_name="sekonder",kademe=sekonder_kademe,values=sekonder_group_list)
    
    insert_kademe_value(ws=ws,start_row=10+primer_number,kademe_name="primer",kademe=primer_kademe,values=primer_group_list)
    
    insert_image_trafo(ws=ws,last_row=max_number+46,image=Image(IMAGE_DIR +"o_mui.png"))
    
    if primer_baglanti=="Seri":
        ws.add_image(Image(IMAGE_DIR + 'seri.jpg'), "G" + str(12))
    elif primer_baglanti=="Paralel":
        ws.add_image(Image(IMAGE_DIR + 'paralel.jpg'), "G" + str(12))
    else:
        print("Bağlantı seçimi hatalı : ",primer_baglanti)
    if sekonder_baglanti=="Seri":

        ws.add_image(Image(IMAGE_DIR + 'seri.jpg'), "O" + str(12))
    elif sekonder_baglanti=="Paralel":
        ws.add_image(Image(IMAGE_DIR + 'paralel.jpg'), "O" + str(12))
    else:
        pass
    ws2=wb.get_sheet_by_name("M_Listesi")
    start_row_list=3
    mlz_listesi_olustur(ws2,start_row_list)
    
    try:
        wb_save(DIR=REPORT_DIR,name='izolasyon_MonoFazUI.xlsx',wb=wb)
        file = REPORT_DIR+"izolasyon_MonoFazUI.xlsx"
        os.startfile(file)
    except Exception as err:
        print(err)    
def monofazSont_printout():
    wb = load_workbook(REPORT_DIR+ 'sont_template.xlsx')
    ws=wb.active
    topline(ws=ws,tarih=datetime.datetime.now(),trafo_tipi=trafo_tipi)
    
    bottomline(ws=ws,last_row=30)
    rightSide_values(ws=ws)
    max_number,primer_number,sek_number,va_number=find_last_row(ws=ws,primer_kademe=primer_kademe,sekonder_kademe=0,va_kademe=0,va_altkademe=0,va_enabled=False)

    insert_row(ws=ws,start_row=30,number=max_number+4)
    ws.merge_cells(start_row=max_number+4+31, start_column=2, end_column=5, end_row=max_number+4+31)
    ws.merge_cells(start_row=max_number+4+32, start_column=2, end_column=5, end_row=max_number+4+32)
    ws.merge_cells(start_row=max_number+4+33, start_column=2, end_column=5, end_row=max_number+4+33)
    ws.merge_cells(start_row=max_number+4+34, start_column=2, end_column=5, end_row=max_number+4+34)
    ws.merge_cells(start_row=max_number+4+35, start_column=2, end_column=5, end_row=max_number+4+35)
    ws.merge_cells(start_row=max_number+4+36, start_column=2, end_column=5, end_row=max_number+4+36)
    start_row=va_number
    
    
    
    insert_kademe_value(ws=ws,start_row=10+primer_number,kademe_name="primer",kademe=primer_kademe,values=primer_group_list)
    
    insert_image_trafo(ws=ws,last_row=max_number+45,image=Image(IMAGE_DIR +"o_mui.png"))
    
    
    ws2=wb.get_sheet_by_name("M_Listesi")
    start_row_list=3
    mlz_listesi_olustur(ws2,start_row_list)
    
    try:
        wb_save(DIR=REPORT_DIR,name='MonoFazSont.xlsx',wb=wb)
        file = REPORT_DIR+"MonoFazSont.xlsx"
        os.startfile(file)
    except Exception as err:
        print(err)    
def trifazSont_printout():
    wb = load_workbook(REPORT_DIR+ 'sont_template.xlsx')
    ws=wb.active
    topline(ws=ws,tarih=datetime.datetime.now(),trafo_tipi=trafo_tipi)
    
    bottomline(ws=ws,last_row=30)
    rightSide_values(ws=ws)
    max_number,primer_number,sek_number,va_number=find_last_row(ws=ws,primer_kademe=primer_kademe,sekonder_kademe=0,va_kademe=0,va_altkademe=0,va_enabled=False)

    insert_row(ws=ws,start_row=30,number=max_number+4)
    ws.merge_cells(start_row=max_number+4+31, start_column=2, end_column=5, end_row=max_number+4+31)
    ws.merge_cells(start_row=max_number+4+32, start_column=2, end_column=5, end_row=max_number+4+32)
    ws.merge_cells(start_row=max_number+4+33, start_column=2, end_column=5, end_row=max_number+4+33)
    ws.merge_cells(start_row=max_number+4+34, start_column=2, end_column=5, end_row=max_number+4+34)
    ws.merge_cells(start_row=max_number+4+35, start_column=2, end_column=5, end_row=max_number+4+35)
    ws.merge_cells(start_row=max_number+4+36, start_column=2, end_column=5, end_row=max_number+4+36)
    start_row=va_number
    
    
    
    insert_kademe_value(ws=ws,start_row=10+primer_number,kademe_name="primer",kademe=primer_kademe,values=primer_group_list)
    
    insert_image_trafo(ws=ws,last_row=max_number+45,image=Image(IMAGE_DIR +"o_mui.png"))
    
    
    ws2=wb.get_sheet_by_name("M_Listesi")
    start_row_list=3
    mlz_listesi_olustur(ws2,start_row_list)
    
    try:
        wb_save(DIR=REPORT_DIR,name='MonoFazSont.xlsx',wb=wb)
        file = REPORT_DIR+"MonoFazSont.xlsx"
        os.startfile(file)
    except Exception as err:
        print(err)    
def monofazSok_printout():
    wb = load_workbook(REPORT_DIR+ 'sont_template.xlsx')
    ws=wb.active
    topline(ws=ws,tarih=datetime.datetime.now(),trafo_tipi=trafo_tipi)
    
    bottomline(ws=ws,last_row=30)
    rightSide_values(ws=ws)
    max_number,primer_number,sek_number,va_number=find_last_row(ws=ws,primer_kademe=primer_kademe,sekonder_kademe=0,va_kademe=0,va_altkademe=0,va_enabled=False)

    insert_row(ws=ws,start_row=30,number=max_number+4)
    ws.merge_cells(start_row=max_number+4+31, start_column=2, end_column=5, end_row=max_number+4+31)
    ws.merge_cells(start_row=max_number+4+32, start_column=2, end_column=5, end_row=max_number+4+32)
    ws.merge_cells(start_row=max_number+4+33, start_column=2, end_column=5, end_row=max_number+4+33)
    ws.merge_cells(start_row=max_number+4+34, start_column=2, end_column=5, end_row=max_number+4+34)
    ws.merge_cells(start_row=max_number+4+35, start_column=2, end_column=5, end_row=max_number+4+35)
    ws.merge_cells(start_row=max_number+4+36, start_column=2, end_column=5, end_row=max_number+4+36)
    start_row=va_number
    
    
    
    insert_kademe_value(ws=ws,start_row=10+primer_number,kademe_name="primer",kademe=primer_kademe,values=primer_group_list)
    
    insert_image_trafo(ws=ws,last_row=max_number+45,image=Image(IMAGE_DIR +"o_mui.png"))
    
    
    ws2=wb.get_sheet_by_name("M_Listesi")
    start_row_list=3
    mlz_listesi_olustur(ws2,start_row_list)
    
    try:
        wb_save(DIR=REPORT_DIR,name='MonoFazSok.xlsx',wb=wb)
        file = REPORT_DIR+"MonoFazSok.xlsx"
        os.startfile(file)
    except Exception as err:
        print(err)    
def trifazSok_printout():
    wb = load_workbook(REPORT_DIR+ 'sont_template.xlsx')
    ws=wb.active
    topline(ws=ws,tarih=datetime.datetime.now(),trafo_tipi=trafo_tipi)
    
    bottomline(ws=ws,last_row=30)
    rightSide_values(ws=ws)
    max_number,primer_number,sek_number,va_number=find_last_row(ws=ws,primer_kademe=primer_kademe,sekonder_kademe=0,va_kademe=0,va_altkademe=0,va_enabled=False)

    insert_row(ws=ws,start_row=30,number=max_number+4)
    ws.merge_cells(start_row=max_number+4+31, start_column=2, end_column=5, end_row=max_number+4+31)
    ws.merge_cells(start_row=max_number+4+32, start_column=2, end_column=5, end_row=max_number+4+32)
    ws.merge_cells(start_row=max_number+4+33, start_column=2, end_column=5, end_row=max_number+4+33)
    ws.merge_cells(start_row=max_number+4+34, start_column=2, end_column=5, end_row=max_number+4+34)
    ws.merge_cells(start_row=max_number+4+35, start_column=2, end_column=5, end_row=max_number+4+35)
    ws.merge_cells(start_row=max_number+4+36, start_column=2, end_column=5, end_row=max_number+4+36)
    start_row=va_number
    
    
    
    insert_kademe_value(ws=ws,start_row=10+primer_number,kademe_name="primer",kademe=primer_kademe,values=primer_group_list)
    
    insert_image_trafo(ws=ws,last_row=max_number+45,image=Image(IMAGE_DIR +"o_mui.png"))
    
    
    ws2=wb.get_sheet_by_name("M_Listesi")
    start_row_list=3
    mlz_listesi_olustur(ws2,start_row_list)
    
    try:
        wb_save(DIR=REPORT_DIR,name='TriFazSok.xlsx',wb=wb)
        file = REPORT_DIR+"TriFazSok.xlsx"
        os.startfile(file)
    except Exception as err:
        print(err)    

if __name__ == '__main__':
    pass    
#izolasyon_mono_printout()
    

